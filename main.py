from datetime import datetime, timedelta

from PyQt5.QtSql import QSqlDatabase, QSqlTableModel, QSqlQuery
from PyQt5.QtWidgets import QWidget, QApplication, QMessageBox, QFileDialog, QMainWindow, QSizePolicy
from openpyxl import load_workbook
from PyQt5.QtWidgets import QComboBox, QCompleter
from PyQt5.QtCore import QSortFilterProxyModel, Qt, QDate
import string
import sys

from PyQt5 import QtCore, QtGui, QtWidgets


def daterange(start_date, end_date):  # функция возвращает генератор дат в диапазоне данных
    for n in range(int((end_date - start_date).days)):
        yield start_date + timedelta(n)


class GroupDisplayModel(QtCore.QAbstractTableModel):  # Самодельная модель для отображения расписания в приложении
    def __init__(self, data):
        super(GroupDisplayModel, self).__init__()
        self.horizontalHeaders = [''] * 6
        headers = ["Группа", "Дисциплина", "Место проведения", "Дата проведения", "Время начала", "Время окончания"]
        for i, name in enumerate(headers):
            self.setHeaderData(i, Qt.Horizontal, name)

        self.data_ = data

    def setHeaderData(self, section, orientation, data, role=Qt.EditRole):  # для заголовков
        if orientation == Qt.Horizontal and role in (Qt.DisplayRole, Qt.EditRole):
            try:
                self.horizontalHeaders[section] = data
                return True
            except:
                return False
        return super().setHeaderData(section, orientation, data, role)

    def headerData(self, section, orientation, role=Qt.DisplayRole):  # для заголовков
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            try:
                return self.horizontalHeaders[section]
            except:
                pass
        return super().headerData(section, orientation, role)

    def data(self, index, role):
        if role == Qt.DisplayRole:
            value = self.data_[index.row()][index.column()]

            if isinstance(value, float):
                return "%.2f" % value

            return value

    def rowCount(self, index):
        return len(self.data_)

    def columnCount(self, index):
        return len(self.data_[0])


class MedSchedule(QMainWindow):  # главное окно с расписанием
    # noinspection PyUnresolvedReferences
    def __init__(self):
        super().__init__()
        self.init_DB()
        self.QTdb = QSqlDatabase.addDatabase('QSQLITE')
        self.QTdb.setDatabaseName("Master.sqlite")
        self.QTdb.open()
        self.editor = DictChange(self.QTdb, parent=self)
        self.setupUi(self)
        self.action.triggered.connect(self.showEditor)
        self.pB_Plus.clicked.connect(self.addElement)
        self.pB_Minus.clicked.connect(self.delElement)
        self.DisplayModel = GroupDisplayModel(self.get_schedule())
        self.model = QSqlTableModel(self, self.QTdb)
        self.model.setTable('schedule')
        self.tV.setModel(self.DisplayModel)
        self.adder = ScheduleEditor(self.model, parent=self)

    def update_display(self):  # обновление модели отображения после внесения изменений
        self.DisplayModel = GroupDisplayModel(self.get_schedule())
        self.tV.setModel(self.DisplayModel)

    def get_schedule(self):  # парсинг из таблицы schedule
        query = QSqlQuery(self.QTdb)
        query.exec(f"""SELECT COUNT(*) FROM schedule """)
        query.first()
        out = list()
        count = query.value(0)
        query.exec(f"""SELECT * FROM schedule """)
        query.first()
        out.append([query.value(i) for i in range(1, 7)])
        for _ in range(1, count):
            query.next()
            out.append([query.value(i) for i in range(1, 7)])
        return out

    @staticmethod
    def init_DB():  # ВАЖНАЯ ХРЕНЬ!! инициализирует структуру бд при отсутствии таковой
        import sqlite3
        con = sqlite3.connect("Master.sqlite")
        cur = con.cursor()

        auditorium = """CREATE TABLE IF NOT EXISTS rooms("id" INTEGER  PRIMARY KEY AUTOINCREMENT UNIQUE, "name" TEXT, 
            "address" TEXT)"""
        groups = """CREATE TABLE IF NOT EXISTS "groups"("id" INTEGER  PRIMARY KEY AUTOINCREMENT UNIQUE,"name" TEXT, 
            "direction"	TEXT, 
            "course" INTEGER )"""
        schedule = """CREATE TABLE IF NOT EXISTS "schedule"("id" INTEGER  PRIMARY KEY AUTOINCREMENT UNIQUE,"group"	TEXT , 
            "subject" TEXT , 
            "venue" TEXT , 
            "date" TEXT ,  
            "time_start" TEXT , 
            "time_end" TEXT )"""
        subject = """CREATE TABLE IF NOT EXISTS "subjects"("id" INTEGER  PRIMARY KEY AUTOINCREMENT UNIQUE,"name" TEXT , 
            "teacher" TEXT )"""

        cur.execute(auditorium)
        cur.execute(groups)
        cur.execute(schedule)
        cur.execute(subject)
        con.commit()

        con.close()

    def closeEvent(self, event):  # переписанный встроенный ивент закрытия, добавлен вопрос
        reply = QMessageBox.question(self, 'Закрыть', 'Закрыть редактор расписания и сохранить изменения?',
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            event.accept()
            self.adder.close()
        else:
            event.ignore()

    def setupUi(self, MainWindow):  # тут UI, тут всё понятно
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(460, 390)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        spacerItem = QtWidgets.QSpacerItem(238, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout.addItem(spacerItem)
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setText("")
        self.pushButton.setObjectName("pushButton")
        self.horizontalLayout.addWidget(self.pushButton)
        self.gridLayout_2.addLayout(self.horizontalLayout, 0, 0, 1, 3)
        self.scrollArea = QtWidgets.QScrollArea(self.centralwidget)
        self.scrollArea.setWidgetResizable(True)
        self.scrollArea.setObjectName("scrollArea")
        self.sAWC = QtWidgets.QWidget()
        self.sAWC.setGeometry(QtCore.QRect(0, 0, 440, 266))
        self.sAWC.setObjectName("sAWC")
        self.gridLayout = QtWidgets.QGridLayout(self.sAWC)
        self.gridLayout.setObjectName("gridLayout")
        self.tV = QtWidgets.QTableView(self.sAWC)
        self.tV.setMouseTracking(False)
        self.tV.setAutoScrollMargin(18)
        self.tV.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.tV.setObjectName("tV")
        self.gridLayout.addWidget(self.tV, 0, 0, 1, 1)
        self.scrollArea.setWidget(self.sAWC)
        self.gridLayout_2.addWidget(self.scrollArea, 1, 0, 1, 3)
        self.pB_Plus = QtWidgets.QPushButton(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pB_Plus.sizePolicy().hasHeightForWidth())
        self.pB_Plus.setSizePolicy(sizePolicy)
        self.pB_Plus.setText("")
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("./ui/AddIcon.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pB_Plus.setIcon(icon)
        self.pB_Plus.setObjectName("pB_Plus")
        self.gridLayout_2.addWidget(self.pB_Plus, 2, 0, 1, 1)
        self.pB_Minus = QtWidgets.QPushButton(self.centralwidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pB_Minus.sizePolicy().hasHeightForWidth())
        self.pB_Minus.setSizePolicy(sizePolicy)
        self.pB_Minus.setText("")
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap("./ui/DelIcon.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pB_Minus.setIcon(icon1)
        self.pB_Minus.setObjectName("pB_Minus")
        self.gridLayout_2.addWidget(self.pB_Minus, 2, 1, 1, 1)
        spacerItem1 = QtWidgets.QSpacerItem(371, 21, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_2.addItem(spacerItem1, 2, 2, 1, 1)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setNativeMenuBar(False)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 460, 22))
        self.menubar.setObjectName("menubar")
        self.menu = QtWidgets.QMenu(self.menubar)
        self.menu.setObjectName("menu")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.action = QtWidgets.QAction(MainWindow)
        self.action.setObjectName("action")
        self.menu.addAction(self.action)
        self.menubar.addAction(self.menu.menuAction())

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.menu.setTitle(_translate("MainWindow", "Справочники"))
        self.action.setText(_translate("MainWindow", "Редактировать"))

    def get_info(self,
                 table):  # функция собирает ВСЕ данные из данной таблицы (для подгрузки в форму добавления) (только для rooms, groups, subjects)
        return self.editor.get_info(table)

    def showEditor(self):
        self.editor.show()
        self.adder.close()

    def addElement(self):
        self.adder = ScheduleEditor(self.model, parent=self)
        self.adder.show()

    def delElement(self):
        rows = list(set([el.row() for el in self.tV.selectionModel().selectedIndexes()]))
        if rows:
            ask = QMessageBox
            status = ask.question(self, '', 'Вы уверены?', ask.Yes | ask.No)

            if status == ask.Yes:
                for i in rows:
                    self.model.deleteRowFromTable(i)
                self.model.submitAll()
                self.model.clear()
                self.model.setTable('schedule')
                self.model.select()
        self.update_display()


class DictChange(QWidget):  # Редактор Справочников
    smodel: QSqlTableModel
    subjectsName: str
    gmodel: QSqlTableModel
    groupsName: str
    rmodel: QSqlTableModel
    roomsName: str

    # noinspection PyUnresolvedReferences
    def __init__(self, db, parent=None, ):
        self.parent = parent
        super(DictChange, self).__init__()
        self.columns = {'rooms': ['name', 'address'],
                        'groups': ['name', 'direction', 'course'],
                        'subjects': ['name', 'teacher']}
        self.db = db
        self.setupUI(self)

        obj_list = [[self.rmodel, self.tv_Rooms, self.roomsName], [self.gmodel, self.tv_Groups, self.groupsName],
                    [self.smodel, self.tv_Subjects, self.subjectsName]]
        self.tb_AddRoom.clicked.connect(lambda: self.addRow(obj_list[0]))
        self.tb_AddGroup.clicked.connect(lambda: self.addRow(obj_list[1]))
        self.tb_AddSubject.clicked.connect(lambda: self.addRow(obj_list[2]))

        self.tb_DelRoom.clicked.connect(lambda: self.delRow(obj_list[0]))
        self.tb_DelGroup.clicked.connect(lambda: self.delRow(obj_list[1]))
        self.tb_DelSubject.clicked.connect(lambda: self.delRow(obj_list[2]))

        self.pb_ImportRooms.clicked.connect(lambda: self.load(obj_list[0]))
        self.pb_ImportGroups.clicked.connect(lambda: self.load(obj_list[1]))
        self.pb_ImportSubjects.clicked.connect(lambda: self.load(obj_list[2]))

    def load_models(self):  # создаёт модели для таблиц, связывает их с таблицами в БД
        self.roomsName = 'rooms'  # loading rooms
        self.rmodel = QSqlTableModel(self, self.db)
        self.rmodel.setTable(self.roomsName)
        self.rmodel.setEditStrategy(QSqlTableModel.OnFieldChange)
        self.rmodel.select()

        self.groupsName = 'groups'  # loading groups
        self.gmodel = QSqlTableModel(self, self.db)
        self.gmodel.setTable(self.groupsName)
        self.gmodel.setEditStrategy(QSqlTableModel.OnFieldChange)
        self.gmodel.select()

        self.subjectsName = 'subjects'  # loading subjects
        self.smodel = QSqlTableModel(self, self.db)
        self.smodel.setTable(self.subjectsName)
        self.smodel.setEditStrategy(QSqlTableModel.OnFieldChange)
        self.smodel.select()

    def setupUI(self, Form):
        Form.setObjectName("Form")
        Form.resize(917, 585)
        self.load_models()
        self.gridLayout_4 = QtWidgets.QGridLayout(Form)
        self.gridLayout_4.setObjectName("gridLayout_4")
        self.tabWidget = QtWidgets.QTabWidget(Form)
        self.tabWidget.setObjectName("tabWidget")
        self.Groups = QtWidgets.QWidget()
        self.Groups.setObjectName("Groups")
        self.gridLayout = QtWidgets.QGridLayout(self.Groups)
        self.gridLayout.setObjectName("gridLayout")
        self.tv_Groups = QtWidgets.QTableView(self.Groups)
        self.tv_Groups.setObjectName("tv_Groups")
        self.tv_Groups.setModel(self.gmodel)
        self.tv_Groups.hideColumn(0)
        self.gridLayout.addWidget(self.tv_Groups, 0, 0, 1, 1)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.tb_AddGroup = QtWidgets.QToolButton(self.Groups)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("./ui/AddIcon.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.tb_AddGroup.setIcon(icon)
        self.tb_AddGroup.setObjectName("tb_AddGroup")
        self.horizontalLayout.addWidget(self.tb_AddGroup)
        self.tb_DelGroup = QtWidgets.QToolButton(self.Groups)
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap("./ui/DelIcon.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.tb_DelGroup.setIcon(icon1)
        self.tb_DelGroup.setObjectName("tb_DelGroup")
        self.horizontalLayout.addWidget(self.tb_DelGroup)
        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout.addItem(spacerItem)
        self.pb_ImportGroups = QtWidgets.QPushButton(self.Groups)
        self.pb_ImportGroups.setObjectName("pb_ImportGroups")
        self.horizontalLayout.addWidget(self.pb_ImportGroups)
        self.gridLayout.addLayout(self.horizontalLayout, 1, 0, 1, 1)
        self.tabWidget.addTab(self.Groups, "")
        self.Subjects = QtWidgets.QWidget()
        self.Subjects.setObjectName("Subjects")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.Subjects)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.tv_Subjects = QtWidgets.QTableView(self.Subjects)
        self.tv_Subjects.setObjectName("tv_Subjects")
        self.tv_Subjects.setModel(self.smodel)
        self.tv_Subjects.hideColumn(0)
        self.gridLayout_2.addWidget(self.tv_Subjects, 0, 0, 1, 1)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.tb_AddSubject = QtWidgets.QToolButton(self.Subjects)
        self.tb_AddSubject.setIcon(icon)
        self.tb_AddSubject.setObjectName("tb_AddSubject")
        self.horizontalLayout_2.addWidget(self.tb_AddSubject)
        self.tb_DelSubject = QtWidgets.QToolButton(self.Subjects)
        self.tb_DelSubject.setIcon(icon1)
        self.tb_DelSubject.setObjectName("tb_DelSubject")
        self.horizontalLayout_2.addWidget(self.tb_DelSubject)
        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_2.addItem(spacerItem1)
        self.pb_ImportSubjects = QtWidgets.QPushButton(self.Subjects)
        self.pb_ImportSubjects.setObjectName("pb_ImportSubjects")
        self.horizontalLayout_2.addWidget(self.pb_ImportSubjects)
        self.gridLayout_2.addLayout(self.horizontalLayout_2, 1, 0, 1, 1)
        self.tabWidget.addTab(self.Subjects, "")
        self.Rooms = QtWidgets.QWidget()
        self.Rooms.setObjectName("Rooms")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.Rooms)
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.tv_Rooms = QtWidgets.QTableView(self.Rooms)
        self.tv_Rooms.setObjectName("tv_Rooms")
        self.tv_Rooms.setModel(self.rmodel)
        self.tv_Rooms.hideColumn(0)
        self.gridLayout_3.addWidget(self.tv_Rooms, 0, 0, 1, 1)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.tb_AddRoom = QtWidgets.QToolButton(self.Rooms)
        self.tb_AddRoom.setIcon(icon)
        self.tb_AddRoom.setObjectName("tb_AddRoom")
        self.horizontalLayout_3.addWidget(self.tb_AddRoom)
        self.tb_DelRoom = QtWidgets.QToolButton(self.Rooms)
        self.tb_DelRoom.setIcon(icon1)
        self.tb_DelRoom.setObjectName("tb_DelRoom")
        self.horizontalLayout_3.addWidget(self.tb_DelRoom)
        spacerItem2 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_3.addItem(spacerItem2)
        self.pb_ImportRooms = QtWidgets.QPushButton(self.Rooms)
        self.pb_ImportRooms.setObjectName("pb_ImportRooms")
        self.horizontalLayout_3.addWidget(self.pb_ImportRooms)
        self.gridLayout_3.addLayout(self.horizontalLayout_3, 1, 0, 1, 1)
        self.tabWidget.addTab(self.Rooms, "")
        self.gridLayout_4.addWidget(self.tabWidget, 0, 0, 1, 1)

        self.retranslateUi(Form)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def closeEvent(self, event):  # вопрос + проверка на незаполненные ячейки
        reply = QMessageBox.No
        query = QSqlQuery(self.db)

        query.exec("SELECT RowNum from "
                   "(SELECT ROW_NUMBER () OVER (ORDER BY id) RowNum, name, direction, course FROM groups)"
                   "WHERE name IS NULL or direction IS NULL or course IS NULL or name = '' or direction = '' or "
                   "course = ''")
        if query.first():
            groupsNULL = query.value(0)
        else:
            groupsNULL = None
        if groupsNULL is not None:
            alert = QMessageBox.information(self, 'Ошибка сохранения', 'Остались незаполненные данные')
            self.tv_Groups.selectRow(groupsNULL - 1)
            self.tabWidget.setCurrentIndex(0)

        query.exec("SELECT RowNum from "
                   "(SELECT ROW_NUMBER () OVER (ORDER BY id) RowNum, name, teacher FROM subjects)"
                   " WHERE name IS NULL or teacher IS NULL or name = '' or teacher = ''")
        if query.first():
            subjNULL = query.value(0)
        else:
            subjNULL = None
        if subjNULL is not None:
            alert = QMessageBox.information(self, 'Ошибка сохранения', 'Остались незаполненные данные')
            self.tv_Subjects.selectRow(subjNULL - 1)
            self.tabWidget.setCurrentIndex(1)
        query.exec("SELECT RowNum from "
                   "(SELECT ROW_NUMBER () OVER (ORDER BY id) RowNum, name, address FROM rooms)"
                   " WHERE name IS NULL or address IS NULL or name = '' or address = ''")
        if query.first():
            roomsNULL = query.value(0)
        else:
            roomsNULL = None
        if roomsNULL is not None:
            alert = QMessageBox.information(self, 'Ошибка сохранения', 'Остались незаполненные данные')
            self.tv_Rooms.selectRow(roomsNULL - 1)
            self.tabWidget.setCurrentIndex(2)
        if not any([groupsNULL, subjNULL, roomsNULL]):
            reply = QMessageBox.question(self, 'Закрыть', 'Закрыть редактор справочников и сохранить изменения?',
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            event.accept()

        else:
            event.ignore()

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Form"))
        self.tb_AddGroup.setText(_translate("Form", "..."))
        self.tb_DelGroup.setText(_translate("Form", "..."))
        self.pb_ImportGroups.setText(_translate("Form", "Import from Excel"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.Groups), _translate("Form", "Учебные группы"))
        self.tb_AddSubject.setText(_translate("Form", "..."))
        self.tb_DelSubject.setText(_translate("Form", "..."))
        self.pb_ImportSubjects.setText(_translate("Form", "Import from Excel"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.Subjects), _translate("Form", "Дисциплины"))
        self.tb_AddRoom.setText(_translate("Form", "..."))
        self.tb_DelRoom.setText(_translate("Form", "..."))
        self.pb_ImportRooms.setText(_translate("Form", "Import from Excel"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.Rooms), _translate("Form", "Аудитории"))

    @staticmethod
    def addRow(toAdd):
        record = toAdd[0].record()
        toAdd[0].insertRecord(-1, record)
        toAdd[0].submitAll()
        toAdd[0].clear()
        toAdd[0].setTable(toAdd[2])
        toAdd[0].select()
        toAdd[1].selectRow(toAdd[1].model().rowCount() - 1)
        toAdd[1].hideColumn(0)

    def delRow(self, toDel):
        rows = list(set([el.row() for el in toDel[1].selectionModel().selectedIndexes()]))
        if rows:
            ask = QMessageBox
            status = ask.question(self, '', 'Вы уверены?', ask.Yes | ask.No)

            if status == ask.Yes:
                for i in rows:
                    toDel[0].deleteRowFromTable(i)
                toDel[0].submitAll()
                toDel[0].clear()
                toDel[0].setTable(toDel[2])
                toDel[0].select()
                toDel[1].selectRow(rows[0] - 1)
        toDel[1].hideColumn(0)

    def load(self, toLoadInto):  # подгрузка из .xls .xlsx
        file, status = QFileDialog.getOpenFileName()
        if status:
            record = toLoadInto[0].record()
            wb = load_workbook(file)
            ws1 = wb['Лист1']
            rc = ws1.max_row
            cc = ws1.max_column
            column_names = list(string.ascii_uppercase)
            record.remove(record.indexOf("id"))
            for i in range(1, rc + 1):
                for j in range(len(self.columns[toLoadInto[2]])):
                    target = column_names[j] + str(i)
                    record.setValue(self.columns[toLoadInto[2]][j], ws1[target].value)
                toLoadInto[0].insertRecord(-1, record)
                toLoadInto[0].submitAll()

    def get_info(self, table):  # возвращает ВСЕ данные из данной таблицы (только для rooms, groups, subjects)
        lengths = {'rooms': 2,
                   'groups': 3,
                   'subjects': 2}
        out = list()
        query = QSqlQuery(self.db)
        query.exec(f"SELECT COUNT(*) FROM {table}")
        print(query.first())
        count = query.value(0)
        query.exec(f'SELECT * FROM {table}')
        query.first()
        out.append([query.value(i) for i in range(1, lengths[table] + 1)])
        for _ in range(1, count):
            query.next()
            out.append([query.value(i) for i in range(1, lengths[table] + 1)])
        return out


# class ExtendedCombo(QComboBox):     НАРАБОТКА УЛУЧШЕННОГО ComboBox, было в планах, не пригодилось
#     # noinspection PyUnresolvedReferences
#     def __init__(self, parent=None):
#         super(ExtendedCombo, self).__init__(parent)
#
#         self.setFocusPolicy(Qt.StrongFocus)
#         self.setEditable(True)
#         self.completer = QCompleter(self)
#
#         # always show all completions
#         self.completer.setCompletionMode(QCompleter.UnfilteredPopupCompletion)
#         self.pFilterModel = QSortFilterProxyModel(self)
#         self.pFilterModel.setFilterCaseSensitivity(Qt.CaseInsensitive)
#
#         self.completer.setPopup(self.view())
#
#         self.setCompleter(self.completer)
#
#         self.lineEdit().textEdited.connect(self.pFilterModel.setFilterFixedString)
#         self.completer.activated.connect(self.setTextIfCompleterIsClicked)
#
#     def setModel(self, model):
#         super(ExtendedCombo, self).setModel(model)
#         self.pFilterModel.setSourceModel(model)
#         self.completer.setModel(self.pFilterModel)
#
#     def setModelColumn(self, column):
#         self.completer.setCompletionColumn(column)
#         self.pFilterModel.setFilterKeyColumn(column)
#         super(ExtendedCombo, self).setModelColumn(column)
#
#     def view(self):
#         return self.completer.popup()
#
#     def index(self):
#         return self.currentIndex()
#
#     def setTextIfCompleterIsClicked(self, text):
#         if text:
#             index = self.findText(text)
#             self.setCurrentIndex(index)
#

class ScheduleEditor(QWidget):  # форма добавления записей в расписание
    model: QSqlTableModel

    # noinspection PyUnresolvedReferences
    def __init__(self, schModel, parent=None):
        self.model = schModel
        self.parent = parent
        super(ScheduleEditor, self).__init__()
        self.setupUi(self)
        self.flag = self.rB_Repeat.isChecked()
        self.non_repeating = [self.lbl_Date, self.dE_Single]
        self.chB_DotW = [self.chB_Mo, self.chB_Tu, self.chB_We, self.chB_Th, self.chB_Fr, self.chB_Sa]
        self.repeating = [self.lbl_DotW, self.lbl_Mo, self.lbl_Tu, self.lbl_We, self.lbl_Th, self.lbl_Fr, self.lbl_Sa,
                          self.lbl_Dates, self.lbl_DashBD, self.dE_RepeatStart, self.dE_RepeatEnd] + self.chB_DotW
        self.rB_Single.clicked.connect(self.repeat_choice)
        self.rB_Repeat.clicked.connect(self.repeat_choice)
        self.bB.button(QtWidgets.QDialogButtonBox.Cancel).clicked.connect(self.close)
        self.bB.button(QtWidgets.QDialogButtonBox.Ok).clicked.connect(self.submit)

    def submit(self):  # собирает данные с UI, добавляет их построчно в таблицу
        if self.flag:
            if any([day.isChecked() for day in self.chB_DotW]):
                days = [i for i, el in enumerate([day.isChecked() for day in self.chB_DotW]) if el]
                date_start = self.dE_RepeatStart.date().toPyDate()
                date_end = self.dE_RepeatEnd.date().toPyDate()
                time_start = self.tE_Start.time().toPyTime()
                time_end = self.tE_End.time().toPyTime()
                if time_end < time_start:
                    info = QMessageBox.information(self, 'Ошибка добавления', 'Неверное время занятия.')
                    return 0
                print(time_end < time_start)
            else:
                info = QMessageBox.information(self, 'Ошибка добавления', 'Выберите дни недели.')
                return 0
        else:
            date_start = self.dE_Single.date().toPyDate()
            date_end = None
            time_start = self.tE_Start.time().toPyTime()
            time_end = self.tE_End.time().toPyTime()
            if time_end < time_start:
                info = QMessageBox.information(self, 'Ошибка добавления', 'Неверное время занятия.')
                return 0
        if date_end:
            if date_end < date_start:
                info = QMessageBox.information(self, 'Ошибка добавления', 'Неверные даты проведения занятия.')
                return 0
        group = self.cB_Group.currentText()
        subject = self.cB_Subject.currentText()
        venue = self.cB_Venue.currentText()
        record = self.model.record()
        if self.check_intersections(group, subject, venue, date_start, time_start, time_end):
            info = QMessageBox.information(self, 'Ошибка добавления', 'Неверные данные.')
            return 0
        else:
            toBeAdded = list()
            if not self.flag:
                record.setValue('date', date_start)
                toBeAdded.append([group, subject, venue, date_start, time_start, time_end])
            else:
                dates = [i for i in daterange(date_start, date_end) if i.weekday() in days]
                for date in dates:
                    toBeAdded.append([group, subject, venue, date, time_start, time_end])
            for el in toBeAdded:
                record.remove(record.indexOf('id'))
                record.setValue('group', el[0])
                record.setValue('subject', el[1])
                record.setValue('venue', el[2])
                record.setValue('date', str(el[3]))
                record.setValue('time_start', str(el[4]))
                record.setValue('time_end', str(el[5]))
                self.model.insertRecord(-1, record)
            self.model.submitAll()
            self.parent.update_display()

    def check_intersections(self, group, subject, venue, date, time_start, time_end):  # проверка на пересечения
        query = QSqlQuery(self.parent.QTdb)
        query.exec("SELECT COUNT(*) FROM schedule")
        print(query.first(), "checks")
        count = query.value(0)
        if count:
            query.exec(f"""SELECT * FROM schedule WHERE "group" = "{group}" """)
            print(query.first(), group)
            out = list()
            out.append([query.value(i) for i in range(1, 8)])
            print(out)
            for _ in range(1, count):
                query.next()
                out.append([query.value(i) for i in range(1, 8)])
            if out:
                for el in out:
                    if el[3] == date.strftime("%Y-%m-%d") and self.check_time_intersections(datetime.strptime(el[4], "%H:%M:%S").time(), datetime.strptime(el[5], "%H:%M:%S").time(), time_start, time_end):
                        return True

            query.exec(f"""SELECT * FROM schedule WHERE "venue" = "{venue}" """)
            query.first()
            out = list()
            out.append([query.value(i) for i in range(1, 8)])
            for _ in range(1, count):
                query.next()
                out.append([query.value(i) for i in range(1, 8)])
            if out:
                for el in out:
                    if el[1] == subject:
                        return False
                    elif el[3] == date.strftime("%Y-%m-%d") and self.check_time_intersections(datetime.strptime(el[4], "%H:%M:%S").time(), datetime.strptime(el[5], "%H:%M:%S").time(), time_start, time_end):
                        return True

        return False

    def check_time_intersections(self, time_start, time_end, time_start_to_add, time_end_to_add):
        # проверка на персечения по времени
        if (time_start < time_start_to_add < time_end) or (time_start < time_end_to_add < time_end) or \
                (time_start_to_add <= time_start and time_end_to_add >= time_end):
            return True
        else:
            return False

    def showEvent(self, event):  # подгрузка в ComboBox'ы на открытии формы
        for el in self.parent.get_info('groups'):
            self.cB_Group.addItem(str(el[0]))
        for el in self.parent.get_info('rooms'):
            self.cB_Venue.addItem(str(el[0]))
        for el in self.parent.get_info('subjects'):
            self.cB_Subject.addItem(str(el[0]))
        event.accept()

    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(442, 181)
        self.gridLayout_2 = QtWidgets.QGridLayout(Form)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.splitter_4 = QtWidgets.QSplitter(Form)
        self.splitter_4.setOrientation(QtCore.Qt.Horizontal)
        self.splitter_4.setObjectName("splitter_4")
        self.lbl_Group = QtWidgets.QLabel(self.splitter_4)
        self.lbl_Group.setObjectName("lbl_Group")
        self.cB_Group = QtWidgets.QComboBox(self.splitter_4)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.cB_Group.sizePolicy().hasHeightForWidth())
        self.cB_Group.setSizePolicy(sizePolicy)
        self.cB_Group.setObjectName("cB_Group")
        self.gridLayout_2.addWidget(self.splitter_4, 0, 0, 1, 1)
        self.rB_Single = QtWidgets.QRadioButton(Form)
        self.rB_Single.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.rB_Single.sizePolicy().hasHeightForWidth())
        self.rB_Single.setSizePolicy(sizePolicy)
        self.rB_Single.setAutoFillBackground(False)
        self.rB_Single.setChecked(True)
        self.rB_Single.setObjectName("rB_Single")
        self.BG_repeatChoice = QtWidgets.QButtonGroup(Form)
        self.BG_repeatChoice.setObjectName("BG_repeatChoice")
        self.BG_repeatChoice.addButton(self.rB_Single)
        self.gridLayout_2.addWidget(self.rB_Single, 0, 1, 1, 1)
        self.splitter_3 = QtWidgets.QSplitter(Form)
        self.splitter_3.setOrientation(QtCore.Qt.Horizontal)
        self.splitter_3.setObjectName("splitter_3")
        self.lbl_Subject = QtWidgets.QLabel(self.splitter_3)
        self.lbl_Subject.setObjectName("lbl_Subject")
        self.cB_Subject = QtWidgets.QComboBox(self.splitter_3)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.cB_Subject.sizePolicy().hasHeightForWidth())
        self.cB_Subject.setSizePolicy(sizePolicy)
        self.cB_Subject.setEditable(True)
        self.cB_Subject.setObjectName("cB_Subject")
        self.gridLayout_2.addWidget(self.splitter_3, 1, 0, 1, 1)
        self.splitter_7 = QtWidgets.QSplitter(Form)
        self.splitter_7.setOrientation(QtCore.Qt.Horizontal)
        self.splitter_7.setObjectName("splitter_7")
        self.lbl_Date = QtWidgets.QLabel(self.splitter_7)
        self.lbl_Date.setObjectName("lbl_Date")
        self.dE_Single = QtWidgets.QDateEdit(self.splitter_7)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.dE_Single.sizePolicy().hasHeightForWidth())
        self.dE_Single.setSizePolicy(sizePolicy)
        self.dE_Single.setObjectName("dE_Single")
        self.gridLayout_2.addWidget(self.splitter_7, 1, 1, 1, 1)
        self.splitter_2 = QtWidgets.QSplitter(Form)
        self.splitter_2.setOrientation(QtCore.Qt.Horizontal)
        self.splitter_2.setObjectName("splitter_2")
        self.lbl_Venue = QtWidgets.QLabel(self.splitter_2)
        self.lbl_Venue.setObjectName("lbl_Venue")
        self.cB_Venue = QtWidgets.QComboBox(self.splitter_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.cB_Venue.sizePolicy().hasHeightForWidth())
        self.cB_Venue.setSizePolicy(sizePolicy)
        self.cB_Venue.setEditable(True)
        self.cB_Venue.setObjectName("cB_Venue")
        self.gridLayout_2.addWidget(self.splitter_2, 2, 0, 1, 1)
        self.rB_Repeat = QtWidgets.QRadioButton(Form)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.rB_Repeat.sizePolicy().hasHeightForWidth())
        self.rB_Repeat.setSizePolicy(sizePolicy)
        self.rB_Repeat.setObjectName("rB_Repeat")
        self.BG_repeatChoice.addButton(self.rB_Repeat)
        self.gridLayout_2.addWidget(self.rB_Repeat, 2, 1, 1, 1)
        self.splitter = QtWidgets.QSplitter(Form)
        self.splitter.setOrientation(QtCore.Qt.Horizontal)
        self.splitter.setObjectName("splitter")
        self.lbl_Time = QtWidgets.QLabel(self.splitter)
        self.lbl_Time.setObjectName("lbl_Time")
        self.layoutWidget = QtWidgets.QWidget(self.splitter)
        self.layoutWidget.setObjectName("layoutWidget")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.layoutWidget)
        self.horizontalLayout.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.tE_Start = QtWidgets.QTimeEdit(self.layoutWidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.tE_Start.sizePolicy().hasHeightForWidth())
        self.tE_Start.setSizePolicy(sizePolicy)
        self.tE_Start.setObjectName("tE_Start")
        self.horizontalLayout.addWidget(self.tE_Start)
        self.lbl_DashBT = QtWidgets.QLabel(self.layoutWidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lbl_DashBT.sizePolicy().hasHeightForWidth())
        self.lbl_DashBT.setSizePolicy(sizePolicy)
        self.lbl_DashBT.setObjectName("lbl_DashBT")
        self.horizontalLayout.addWidget(self.lbl_DashBT)
        self.tE_End = QtWidgets.QTimeEdit(self.layoutWidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.tE_End.sizePolicy().hasHeightForWidth())
        self.tE_End.setSizePolicy(sizePolicy)
        self.tE_End.setObjectName("tE_End")
        self.horizontalLayout.addWidget(self.tE_End)
        self.gridLayout_2.addWidget(self.splitter, 3, 0, 1, 1)
        self.splitter_6 = QtWidgets.QSplitter(Form)
        self.splitter_6.setOrientation(QtCore.Qt.Horizontal)
        self.splitter_6.setObjectName("splitter_6")
        self.lbl_DotW = QtWidgets.QLabel(self.splitter_6)
        self.lbl_DotW.setEnabled(False)
        self.lbl_DotW.setObjectName("lbl_DotW")
        self.layoutWidget1 = QtWidgets.QWidget(self.splitter_6)
        self.layoutWidget1.setObjectName("layoutWidget1")
        self.gridLayout = QtWidgets.QGridLayout(self.layoutWidget1)
        self.gridLayout.setContentsMargins(0, 0, 0, 0)
        self.gridLayout.setObjectName("gridLayout")
        self.chB_Mo = QtWidgets.QCheckBox(self.layoutWidget1)
        self.chB_Mo.setEnabled(False)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.chB_Mo.sizePolicy().hasHeightForWidth())
        self.chB_Mo.setSizePolicy(sizePolicy)
        self.chB_Mo.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.chB_Mo.setText("")
        self.chB_Mo.setObjectName("chB_Mo")
        self.gridLayout.addWidget(self.chB_Mo, 0, 0, 1, 1)
        self.chB_Tu = QtWidgets.QCheckBox(self.layoutWidget1)
        self.chB_Tu.setEnabled(False)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.chB_Tu.sizePolicy().hasHeightForWidth())
        self.chB_Tu.setSizePolicy(sizePolicy)
        self.chB_Tu.setText("")
        self.chB_Tu.setObjectName("chB_Tu")
        self.gridLayout.addWidget(self.chB_Tu, 0, 1, 1, 1)
        self.chB_We = QtWidgets.QCheckBox(self.layoutWidget1)
        self.chB_We.setEnabled(False)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.chB_We.sizePolicy().hasHeightForWidth())
        self.chB_We.setSizePolicy(sizePolicy)
        self.chB_We.setText("")
        self.chB_We.setObjectName("chB_We")
        self.gridLayout.addWidget(self.chB_We, 0, 2, 1, 1)
        self.chB_Th = QtWidgets.QCheckBox(self.layoutWidget1)
        self.chB_Th.setEnabled(False)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.chB_Th.sizePolicy().hasHeightForWidth())
        self.chB_Th.setSizePolicy(sizePolicy)
        self.chB_Th.setText("")
        self.chB_Th.setObjectName("chB_Th")
        self.gridLayout.addWidget(self.chB_Th, 0, 3, 1, 1)
        self.chB_Fr = QtWidgets.QCheckBox(self.layoutWidget1)
        self.chB_Fr.setEnabled(False)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.chB_Fr.sizePolicy().hasHeightForWidth())
        self.chB_Fr.setSizePolicy(sizePolicy)
        self.chB_Fr.setText("")
        self.chB_Fr.setObjectName("chB_Fr")
        self.gridLayout.addWidget(self.chB_Fr, 0, 4, 1, 1)
        self.chB_Sa = QtWidgets.QCheckBox(self.layoutWidget1)
        self.chB_Sa.setEnabled(False)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.chB_Sa.sizePolicy().hasHeightForWidth())
        self.chB_Sa.setSizePolicy(sizePolicy)
        self.chB_Sa.setText("")
        self.chB_Sa.setObjectName("chB_Sa")
        self.gridLayout.addWidget(self.chB_Sa, 0, 5, 1, 1)
        self.lbl_Mo = QtWidgets.QLabel(self.layoutWidget1)
        self.lbl_Mo.setEnabled(False)
        self.lbl_Mo.setObjectName("lbl_Mo")
        self.gridLayout.addWidget(self.lbl_Mo, 1, 0, 1, 1)
        self.lbl_Tu = QtWidgets.QLabel(self.layoutWidget1)
        self.lbl_Tu.setEnabled(False)
        self.lbl_Tu.setObjectName("lbl_Tu")
        self.gridLayout.addWidget(self.lbl_Tu, 1, 1, 1, 1)
        self.lbl_We = QtWidgets.QLabel(self.layoutWidget1)
        self.lbl_We.setEnabled(False)
        self.lbl_We.setObjectName("lbl_We")
        self.gridLayout.addWidget(self.lbl_We, 1, 2, 1, 1)
        self.lbl_Th = QtWidgets.QLabel(self.layoutWidget1)
        self.lbl_Th.setEnabled(False)
        self.lbl_Th.setObjectName("lbl_Th")
        self.gridLayout.addWidget(self.lbl_Th, 1, 3, 1, 1)
        self.lbl_Fr = QtWidgets.QLabel(self.layoutWidget1)
        self.lbl_Fr.setEnabled(False)
        self.lbl_Fr.setObjectName("lbl_Fr")
        self.gridLayout.addWidget(self.lbl_Fr, 1, 4, 1, 1)
        self.lbl_Sa = QtWidgets.QLabel(self.layoutWidget1)
        self.lbl_Sa.setEnabled(False)
        self.lbl_Sa.setObjectName("lbl_Sa")
        self.gridLayout.addWidget(self.lbl_Sa, 1, 5, 1, 1)
        self.gridLayout_2.addWidget(self.splitter_6, 3, 1, 2, 1)
        self.lbl_Space = QtWidgets.QLabel(Form)
        self.lbl_Space.setText("")
        self.lbl_Space.setObjectName("lbl_Space")
        self.gridLayout_2.addWidget(self.lbl_Space, 4, 0, 2, 1)
        self.splitter_5 = QtWidgets.QSplitter(Form)
        self.splitter_5.setOrientation(QtCore.Qt.Horizontal)
        self.splitter_5.setObjectName("splitter_5")
        self.lbl_Dates = QtWidgets.QLabel(self.splitter_5)
        self.lbl_Dates.setEnabled(False)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lbl_Dates.sizePolicy().hasHeightForWidth())
        self.lbl_Dates.setSizePolicy(sizePolicy)
        self.lbl_Dates.setObjectName("lbl_Dates")
        self.layoutWidget2 = QtWidgets.QWidget(self.splitter_5)
        self.layoutWidget2.setObjectName("layoutWidget2")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout(self.layoutWidget2)
        self.horizontalLayout_2.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.dE_RepeatStart = QtWidgets.QDateEdit(self.layoutWidget2)
        self.dE_RepeatStart.setEnabled(False)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.dE_RepeatStart.sizePolicy().hasHeightForWidth())
        self.dE_RepeatStart.setSizePolicy(sizePolicy)
        self.dE_RepeatStart.setObjectName("dE_RepeatStart")
        self.horizontalLayout_2.addWidget(self.dE_RepeatStart)
        self.lbl_DashBD = QtWidgets.QLabel(self.layoutWidget2)
        self.lbl_DashBD.setEnabled(False)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lbl_DashBD.sizePolicy().hasHeightForWidth())
        self.lbl_DashBD.setSizePolicy(sizePolicy)
        self.lbl_DashBD.setObjectName("lbl_DashBD")
        self.horizontalLayout_2.addWidget(self.lbl_DashBD)
        self.dE_RepeatEnd = QtWidgets.QDateEdit(self.layoutWidget2)
        self.dE_RepeatEnd.setEnabled(False)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.dE_RepeatEnd.sizePolicy().hasHeightForWidth())
        self.dE_RepeatEnd.setSizePolicy(sizePolicy)
        self.dE_RepeatEnd.setObjectName("dE_RepeatEnd")
        self.horizontalLayout_2.addWidget(self.dE_RepeatEnd)
        self.gridLayout_2.addWidget(self.splitter_5, 5, 1, 1, 1)
        self.bB = QtWidgets.QDialogButtonBox(Form)
        self.bB.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.bB.setStandardButtons(QtWidgets.QDialogButtonBox.Cancel | QtWidgets.QDialogButtonBox.Ok)
        self.bB.setCenterButtons(False)
        self.bB.setObjectName("bB")
        self.gridLayout_2.addWidget(self.bB, 6, 0, 1, 2)

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Form"))
        self.lbl_Group.setText(_translate("Form", "Учебная группа:"))
        self.rB_Single.setText(_translate("Form", "Не повторяется"))
        self.lbl_Subject.setText(_translate("Form", "Дисциплина:"))
        self.lbl_Date.setText(_translate("Form", "     Дата:"))
        self.lbl_Venue.setText(_translate("Form", "Место проведения:"))
        self.rB_Repeat.setText(_translate("Form", "Повторяется"))
        self.lbl_Time.setText(_translate("Form", "Время:"))
        self.lbl_DashBT.setText(_translate("Form", "—"))
        self.lbl_DotW.setText(_translate("Form", "     День недели:"))
        self.lbl_Mo.setText(_translate("Form", "Пн"))
        self.lbl_Tu.setText(_translate("Form", "Вт"))
        self.lbl_We.setText(_translate("Form", "Ср"))
        self.lbl_Th.setText(_translate("Form", "Чт"))
        self.lbl_Fr.setText(_translate("Form", "Пт"))
        self.lbl_Sa.setText(_translate("Form", "Сб"))
        self.lbl_Dates.setText(_translate("Form", "     Даты:"))
        self.lbl_DashBD.setText(_translate("Form", "—"))

    def repeat_choice(self):  # смена активных элементов UI (повторяющаяся/не повторяющаяся)
        self.flag = self.rB_Repeat.isChecked()
        self.dE_Single.setDate(QDate(2000, 1, 1))
        self.dE_RepeatStart.setDate(QDate(2000, 1, 1))
        self.dE_RepeatEnd.setDate(QDate(2000, 1, 1))  # просто дефолтные даты
        for el in self.chB_DotW:
            el.setChecked(False)
        for el in self.repeating:
            el.setEnabled(self.flag)
        for el in self.non_repeating:
            el.setEnabled(not self.flag)


if __name__ == '__main__':
    app = QApplication(sys.argv)

    mainWindow = MedSchedule()
    mainWindow.show()
    sys.exit(app.exec_())
